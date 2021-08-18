#  MONITORAMENTO URBANO PARA AVALIAÇÃO DE MOBILIDADE URBANA COM 
#        USO DE DRONE   E INTELIGÊNCIA ARTIFICIAL
# -----------------------------------------------------------------
# Programa para multi-detecção e contagem de objetos em uma imagem
# -----------------------------------------------------------------
# PIBIC-EM 2020/2021
# Autores: Giovanna Pavani Martelli e Nícolas Maisonnette Duarte 


# Importando bibliotecas do projeto
import imutils
import cv2  # OpenCV
import numpy as np
from openpyxl import load_workbook, Workbook  # Excel
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font  # Styles do excel
from tkinter import Tk  # Arquivos
from tkinter.filedialog import askopenfilename
from PIL import Image # Imagem

# Retorna a data de captura da imagem do caminho passado por parâmetro
def get_date_taken(path):
    return Image.open(path)._getexif()[36867]

# Declaração de variáveis globais
data = ''
# Pixels de começo e fim das imagens templates
l_xStart = -1
l_yStart = -1
l_xEnd = -1
l_yEnd = -1
# Tamanhos da imagens
l_height = -1
l_width = -1
# Contador de matches
contador = 0
l_valMax = -1
c_valMax = -1
valMax_last = -1
# Vetor de imagens template 
imagensTemplate = [r'.\Project\images#\carUnicoJojo1.png', 
                   r'.\Project\images#\carUnicoJojo2.png',
                   r'.\Project\images#\carUnicoJojo4.png',
                   r'.\Project\images#\carUnicoJojo5.png',]
                  #  r'.\Project\images#\carUnicoJojo7.png',
                  #  r'.\Project\images#\carUnicoJojo8.png',
                  #  r'.\Project\images#\carUnicoJojo9.png',
                  #  r'.\Project\images#\carUnicoJojo10.png',
                  #  r'.\Project\images#\carUnicoJojo11.png',
                  #  r'.\Project\images#\carUnicoJojo12.png',
                   
#vals_maxs = [41602273, 48509457, 66520257, 39139037, 46415409]
#vals_maxs = [42961905, 57359841, 57067197, 1, 1]

# Abre arquivos e imagens
Tk().withdraw()
filename = askopenfilename()

main_image = cv2.imread(filename)
(m_height, m_width) = main_image.shape[:2]
# main_image = cv2.GaussianBlur(main_image, (15, 15), 2) (Tentamos blur, sem sucesso)

# Tela de carregamento
cv2.imshow('Aguarde', cv2.imread(r'.\Project\images#\carregando.png'))
cv2.waitKey(0)

try:
   # Verifica se já existe um arquivo Excel
   workbook = load_workbook("./Project/Resultados.xlsx")
   sheet = workbook.active
except FileNotFoundError:
   # Caso não exista, cria-se um Resultados.xlsx
   workbook = Workbook()
   sheet = workbook.active
   # Setta os headers da tabela
   sheet["A1"].font = Font(bold=True)
   sheet["B1"].font = Font(bold=True)
   sheet["C1"].font = Font(bold=True)
   sheet["A1"] = "Data (y:m:d h:min:s)"
   sheet["B1"] = "Quantidade"
   sheet["C1"] = "Nome do Arquivo"
   sheet.column_dimensions["A"].width = 30
   sheet.column_dimensions["B"].width = 15
   sheet.column_dimensions["C"].width = 120

data = get_date_taken(filename)
i = 0

# Percorre o vetor de templates
while(i < len(imagensTemplate)):
   # Abre o template e armazena o canny
   template = cv2.imread(imagensTemplate[i])
   template = cv2.cvtColor(template, cv2.COLOR_BGR2GRAY)
   template = cv2.Canny(template, 10, 25)
   (height, width) = template.shape[:2]
   
   while True: 
      # gray_imagem = main_image em escala de cinza  
      gray_image = cv2.cvtColor(main_image, cv2.COLOR_BGR2GRAY)
      gray_exib = imutils.resize(gray_image, width = int(500))
      cv2.imshow('Gray', gray_exib)
      temp_found = None

      for scale in np.linspace(0.2, 1.0, 20)[::-1]:
         # Redimensiona a gray_image e armazena o ratio
         resized_img = imutils.resize(gray_image, width = int(gray_image.shape[1] * scale))
         ratio = gray_image.shape[1] / float(resized_img.shape[1])
         if resized_img.shape[0] < height or resized_img.shape[1] < width:
            break
         
         # Converte em edged image para verificar
         e = cv2.Canny(resized_img, 10, 25)
         match = cv2.matchTemplate(e, template, cv2.TM_CCOEFF)
         (_, val_max, _, loc_max) = cv2.minMaxLoc(match)

         # Achou template válido
         if temp_found is None or val_max>temp_found[0]:
            temp_found = (val_max, loc_max, ratio)
            c_valMax = val_max
      
      # ValMax representa, numericamente, a compatibilidade do template com a imagem
      print(f'c_valMax: {c_valMax}') 
      
      if l_valMax >= 0: # Se não é a primeira vez
         if c_valMax == valMax_last: # Verifica repetição do val_max
            contador = contador - 1
            print(f'Quantidade {contador}')
            valMax_last = -1
            break

         valMax_last = c_valMax

         #if c_valMax < vals_maxs[i]:
            #print(f'Quantidade: {contador}')
            #break

      l_valMax = c_valMax
      contador = contador + 1

      # Pega a informação do temp_found e armazena as coordenadas X e Y
      (_, loc_max, r) = temp_found
      (x_start, y_start) = (int(loc_max[0]), int(loc_max[1]))
      (x_end, y_end) = (int((loc_max[0] + width)), int((loc_max[1] + height)))

      l_xStart = x_start
      l_yStart = y_start
      l_xEnd = x_end
      l_yEnd = y_end

      # Desenha uma elipse onde o teplate foi encontrado na main_image
      x_coordinate = (x_end-x_start)/2 + x_start
      y_coordinate = (y_end-y_start)/2 + y_start
      x_coordinate = round(x_coordinate)
      y_coordinate = round(y_coordinate)

      main_image = cv2.ellipse(main_image, (x_coordinate, y_coordinate), (round((x_end - x_start)/1.7), round((y_end - y_start)/1.7)), 0.0, 0.0, 360.0, (255, 0, 0), -1)
      # Redimensiona a main_image para que ela caiba na tela, já que sua resolução original é muito alta
      main_image_resized = imutils.resize(main_image, width = int(500))
      
      #testa uma por uma
      cv2.imshow('Template Found', main_image_resized)
      cv2.waitKey(0)
      
   i = i + 1

local = filename.find('#')
arquivo = filename[local+2:len(filename)]
# Registra resultado no arquivo Excel
sheet.append([data, contador, arquivo])
# Salva
workbook.save(r".\Project\Resultados.xlsx")

# Fecha todas as telas anteriores 
cv2.destroyAllWindows
cv2.waitKey(0)
# Exibe a imagem com as elipses dos templates encontrados
cv2.imshow('Template Found', main_image_resized)
cv2.waitKey(0)
cv2.destroyAllWindows

