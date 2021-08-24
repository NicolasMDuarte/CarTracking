from tkinter import filedialog
import imutils
import os
import cv2  # OpenCV
import numpy as np
from openpyxl import load_workbook, Workbook  # Excel
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font  # Styles do excel
from tkinter import *  # Arquivos
from tkinter.filedialog import askopenfilename
from PIL import Image # Imagem
from PIL import *
from datetime import datetime
from PIL import Image

def imgDate(fn):
    T = None
    try:
        std_fmt = '%Y:%m:%d %H:%M:%S.%f'
        tags = [(36867, 37521),  # (DateTimeOriginal, SubsecTimeOriginal)
                (36868, 37522),  # (DateTimeDigitized, SubsecTimeDigitized)
                (306, 37520), ]  # (DateTime, SubsecTime)
        exif = Image.open(fn)._getexif()
    
        for t in tags:
            dat = exif.get(t[0])
            sub = exif.get(t[1], 0)
    
            # PIL.PILLOW_VERSION >= 3.0 returns a tuple
            dat = dat[0] if type(dat) == tuple else dat
            sub = sub[0] if type(sub) == tuple else sub
            if dat != None: break
    
        if dat == None: return None
        full = '{}.{}'.format(dat, sub)
        T = datetime.strptime(full, std_fmt)
        #T = time.mktime(time.strptime(dat, '%Y:%m:%d %H:%M:%S')) + float('0.%s' % sub)
    except:
        None
    return T

car_classifier = cv2.CascadeClassifier(r'Project\haarcascade_car.xml')
foldername = filedialog.askdirectory()
filenames = os.listdir(foldername)
for filename in filenames:
    qtdCarros = 0
    filename = foldername + "/" + filename
    cap = cv2.imread(filename)
    (height, width) = cap.shape[:2]
    gray = cv2.cvtColor(cap, cv2.COLOR_BGR2GRAY)

    taxa = 35
    parou = False
    while taxa >= 25:
        cars = car_classifier.detectMultiScale(gray, 1.004, taxa) #35
        for (x,y,w,h) in cars:
            qtdCarros = qtdCarros + 1
            cv2.rectangle(cap, (x, y), (x+w, y+h), (0, 0, 255), 2)
            gray_exib = imutils.resize(cap, width = int(1000))
        if(qtdCarros <= 25):
            parou = True
            break
        taxa = taxa - 5

    if(not parou):
        qtdCarros = round(qtdCarros / 3)

    try:
        # Verifica se já existe um arquivo Excel
        workbook = load_workbook("./Project/Resultados.xlsx")
        sheet = workbook.active
    except FileNotFoundError:
        # Caso não exista, cria-se um Resultados.xlsx
        workbook = Workbook()
        sheet = workbook.active
        # Setta os headers da tabela
        sheet["A1"].font = Font(bold=True)
        sheet["B1"].font = Font(bold=True)
        sheet["C1"].font = Font(bold=True)
        sheet["A1"] = "Data (y:m:d h:min:s)"
        sheet["B1"] = "Quantidade"
        sheet["C1"] = "Nome do Arquivo"
        sheet.column_dimensions["A"].width = 30
        sheet.column_dimensions["B"].width = 15
        sheet.column_dimensions["C"].width = 120

    data = imgDate(filename)
    local = filename.find('#')
    arquivo = filename[local+2:len(filename)]
    # Registra resultado no arquivo Excel
    sheet.append([data, qtdCarros, arquivo])
    # Salva
    workbook.save(r".\Project\Resultados.xlsx")
    print(qtdCarros)

#cv2.imshow('Cars', gray_exib)
cv2.waitKey(0)
cv2.destroyAllWindows

#Taxa de erro: ~10% (3 em 32)